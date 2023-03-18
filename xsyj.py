import traceback
from tkinter import *
from tkinter.filedialog import askdirectory, askopenfilename
from tkinter.scrolledtext import ScrolledText

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class Report(Frame):

    def write_excel(self, out_path, month_fund_business):
        for month, fund_busiess in month_fund_business.items():
            wb = Workbook()
            fund_busiess_dict = {}
            for i in range(len(fund_busiess)):
                per_split = fund_busiess[i].split('@')
                fund_code = per_split[8]
                if fund_code in fund_busiess_dict:
                    fund_busiess_dict[fund_code].append(fund_busiess[i])
                else:
                    business = [fund_busiess[i]]
                    fund_busiess_dict[fund_code] = business

            index = 0
            for fund_code, business in fund_busiess_dict.items():
                per_split = business[0].split('@')

                output_sheet = wb.create_sheet(per_split[1], index)
                output_sheet.freeze_panes = 'A4'

                output_sheet.column_dimensions['A'].width = 5.08
                for i in range(2, 11):
                    output_sheet.column_dimensions[get_column_letter(i)].width = 17  # 16.85

                row1 = output_sheet.row_dimensions[1]
                row1.height = 40.5
                cell1_1 = output_sheet.cell(row=1, column=1)  # 第一行第一列的单元格
                cell1_1.font = Font(bold=True, name="黑体", size=16)
                cell1_1.value = per_split[1] + per_split[8] + '收入业绩奖励'  # 单元格赋值
                cell1_1.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                              wrapText=True)  # 居中对齐
                output_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)  # 合并单元格

                row2 = output_sheet.row_dimensions[2]
                row2.height = 25
                cell2_11 = output_sheet.cell(row=2, column=11)
                cell2_11.font = Font(bold=True, name="宋体", size=11)
                cell2_11.value = '单位：元'
                cell2_11.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)

                row3 = output_sheet.row_dimensions[3]
                row3.height = 40.5
                heads = ['序号', '营业部编码', '营业部', '实际销量', '奖励倍数', '基准佣金率', '应到收入\r\n（含税）',
                         '应到收入\r\n（不含税）', '应发业绩奖励\r\n（扣除投保、增值税金及附加）',
                         '实发业绩奖励\r\n（扣除投保、增值税金及附加）', '备注']
                for i in range(0, 11):
                    cell3_i = output_sheet.cell(row=3, column=i + 1)
                    cell3_i.font = Font(bold=True, name="宋体", size=11)
                    cell3_i.value = heads[i]
                    cell3_i.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrapText=True)
                    cell3_i.fill = PatternFill("solid", fgColor="C0C0C0")
                    cell3_i.border = Border(bottom=Side(style='thin', color='000000'),
                                            right=Side(style='thin', color='000000'),
                                            left=Side(style='thin', color='000000'),
                                            top=Side(style='thick', color='000000'))
                for i in range(0, 3):
                    cell3_i = output_sheet.cell(row=3, column=i + 8)
                    cell3_i.fill = PatternFill("solid", fgColor="FF0000")

                req = 1
                for per in business:
                    row = str(req + 3)
                    per_split = per.split('@')
                    income_exclusive_tax = '=D' + row + '*E' + row + '*F' + row
                    income_excluding_tax = '=G' + row + '/1.06'
                    performance_bonus_payable = '=G' + row + '*0.927768*20%'
                    actual_performance_award = '=I' + row
                    content = [req, per_split[2], per_split[3], per_split[4], per_split[5], per_split[6],
                               income_exclusive_tax, income_excluding_tax, performance_bonus_payable,
                               actual_performance_award, '']
                    for i in range(0, 11):
                        cell3_i = output_sheet.cell(row=int(row), column=i + 1)
                        cell3_i.font = Font(bold=False, name='宋体', size=11)
                        cell3_i.border = Border(bottom=Side(style='thin', color='000000'),
                                                right=Side(style='thin', color='000000'),
                                                left=Side(style='thin', color='000000'),
                                                top=Side(style='thin', color='000000'))

                        if i + 1 in (1, 2):
                            cell3_i.value = int(content[i]) if content[i] and content[i] != 'None' else content[i]
                        elif i + 1 in (4, 5):
                            cell3_i.value = float(content[i]) if content[i] and content[i] != 'None' else content[i]
                        elif i + 1 == 6:
                            cell3_i.value = float(content[i]) - 0.0001 if content[i] and content[i] != 'None' else \
                                content[i]
                        else:
                            cell3_i.value = str(content[i]) if content[i] and content[i] != 'None' else content[i]

                        if i + 1 in (4, 5, 7, 8, 9, 10):
                            cell3_i.number_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'  # 设置格式为保留两位小数
                            cell3_i.alignment = openpyxl.styles.Alignment(horizontal='right', vertical='center',
                                                                          wrapText=True)
                        elif i + 1 == 6:
                            cell3_i.number_format = '0.00%'
                            cell3_i.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                          wrapText=True)
                        else:
                            cell3_i.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                          wrapText=True)

                        if i + 1 in (1, 7, 8, 9, 10):
                            cell3_i.font = Font(bold=False, name='宋体', size=11)
                        else:
                            cell3_i.font = Font(bold=True, name='宋体', size=11)

                    req += 1

                # 统计
                tail_row = int(req + 3)
                row3 = output_sheet.row_dimensions[tail_row]
                row3.height = 27.75

                output_sheet.merge_cells(start_row=tail_row, start_column=1, end_row=tail_row, end_column=3)  # 合并单元格

                for i in range(0, 11):
                    tail_cell = output_sheet.cell(row=int(tail_row), column=i + 1)
                    tail_cell.fill = PatternFill("solid", fgColor="C0C0C0")
                    tail_cell.border = Border(bottom=Side(style='thick', color='000000'),
                                              right=Side(style='thin', color='000000'),
                                              left=Side(style='thin', color='000000'),
                                              top=Side(style='thin', color='000000'))
                    tail_cell.number_format = r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)'  # 设置格式为保留两位小数
                    tail_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center',
                                                                    wrapText=True)
                    tail_cell.font = Font(bold=True, name="宋体", size=11)

                    if i + 1 == 1:
                        tail_cell.value = '总计'
                    elif i + 1 == 4:
                        tail_cell.value = '=SUM(D4:D' + str(tail_row - 1) + ')'
                    elif i + 1 == 7:
                        tail_cell.value = '=SUM(G4:G' + str(tail_row - 1) + ')'
                    elif i + 1 == 8:
                        tail_cell.value = '=SUM(H4:H' + str(tail_row - 1) + ')'
                    elif i + 1 == 9:
                        tail_cell.value = '=SUM(I4:I' + str(tail_row - 1) + ')'
                    elif i + 1 == 10:
                        tail_cell.value = '=SUM(J4:J' + str(tail_row - 1) + ')'

            index += 1
            self.run_log_self(out_path + '收入业绩奖励' + month + '.xlsx')
            wb.save(out_path + '\\【收入业绩奖励】' + month + '.xlsx')

        self.run_log_self('解析完成！')

    def read_excel(self, file_name):
        self.run_log_self(file_name)
        wb = load_workbook(file_name, data_only=True)
        ws_b = wb['B-公募基金']
        ws_a = wb['A-实际及考核销量明细']

        fund_name_idx = None
        fund_code_idx = None
        month_idx = None
        commission_rate_idx = None
        fund_company_idx = None
        for column in range(1, ws_b.max_column + 1):
            title = self.__str_wash(ws_b.cell(row=2, column=column).value)
            if title == '基金名称':
                fund_name_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '基金代码':
                fund_code_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '考核系数':
                month_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '佣金率':
                commission_rate_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '基金公司':
                fund_company_idx = column
                self.run_log_self(title + ' ' + str(column))

        income_multiple_idx = None
        for column in range(1, ws_b.max_column + 1):
            title = self.__str_wash(ws_b.cell(row=3, column=column).value)
            if title == '收入倍数':
                income_multiple_idx = column
                self.run_log_self(title + ' ' + str(column))

        month_funds = {}
        for row in ws_b.iter_cols(min_row=4, min_col=fund_code_idx, max_row=ws_b.max_row, max_col=fund_code_idx):
            fund_code_info = None
            for cell in row:
                if not cell.value:
                    month_value = self.__month_format(self.__str_wash(ws_b.cell(row=cell.row, column=month_idx).value))
                    fund_code_info = []
                    month_funds[month_value] = fund_code_info
                else:
                    cell_value_s = self.__str_wash(cell.value)
                    self.run_log_self(cell_value_s)

                    fund_codes = []
                    if '、' in cell_value_s:
                        fund_codes.extend(cell_value_s.split('、'))
                    elif '\\' in cell_value_s:
                        fund_codes.extend(cell_value_s.split('\\'))
                    else:
                        fund_codes.append(cell_value_s)

                    for fund_code in fund_codes:
                        fund_info = fund_code + '@' + self.__str_wash(
                            ws_b.cell(row=cell.row, column=fund_name_idx).value) + '@' + self.__str_wash(
                            ws_b.cell(row=cell.row, column=income_multiple_idx).value) + '@' + self.__str_wash(
                            ws_b.cell(row=cell.row, column=commission_rate_idx).value) + '@' + self.__str_wash(
                            ws_b.cell(row=cell.row, column=fund_company_idx).value) + '@' + self.__str_wash(
                            cell_value_s)
                        fund_code_info.append(fund_info)

        self.run_log_self(month_funds)

        business_code_idx = None
        business_name_idx = None
        business_short_name_idx = None
        for column in range(1, ws_a.max_column + 1):
            title = self.__str_wash(ws_a.cell(row=2, column=column).value)
            if title == '营业部编码':
                business_code_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '营业部全称':
                business_name_idx = column
                self.run_log_self(title + ' ' + str(column))
            elif title == '营业部':
                business_short_name_idx = column
                self.run_log_self(title + ' ' + str(column))

        column_months = {}
        for column in range(1, ws_a.max_column + 1):
            month_value = ws_a.cell(row=1, column=column).value
            if month_value:
                column_months[column] = self.__month_format(self.__str_wash(month_value))
        self.run_log_self(column_months)
        self.run_log_self(column_months.keys())

        month_fund_business = {}
        column_months_keys = list(column_months.keys())
        for i in range(len(column_months_keys)):
            fund_businesses = []
            start_column = column_months_keys[i]
            end_column = column_months_keys[i + 1] if i + 1 < len(column_months_keys) else ws_a.max_column

            idx = start_column
            while idx < end_column and ws_a.cell(row=2, column=idx).value:
                fund_infos = month_funds[column_months[start_column]]
                for fund_info in fund_infos:
                    fund_code = fund_info.split('@')[0]
                    if fund_code in self.__str_wash(ws_a.cell(row=2, column=idx).value):
                        for row in ws_a.iter_cols(min_row=4, min_col=idx, max_row=ws_a.max_row, max_col=idx):
                            for cell in row:
                                business_name = self.__str_wash(ws_a.cell(row=cell.row, column=business_name_idx).value)
                                if str(business_name) == '合计':
                                    break
                                business_short_name = self.__str_wash(
                                    ws_a.cell(row=cell.row, column=business_short_name_idx).value)
                                business_code = self.__str_wash(ws_a.cell(row=cell.row, column=business_code_idx).value)
                                fund_business = str(fund_code) + '@' + str(fund_info.split('@')[1]) + '@' + str(
                                    business_code) + '@' + str(business_name) + '@' + str(cell.value) + '@' + str(
                                    fund_info.split('@')[2]) + '@' + str(fund_info.split('@')[3]) + '@' + str(
                                    fund_info.split('@')[4]) + '@' + str(fund_info.split('@')[5]) + '@' + str(
                                    business_short_name)
                                fund_businesses.append(fund_business)
                                self.run_log_self(fund_business)
                        break
                idx = idx + 1
            month_fund_business[column_months[start_column]] = fund_businesses

        self.run_log_self(month_fund_business)
        return month_fund_business

    @staticmethod
    def __month_format(month_value):
        return month_value.replace('十二', '12') \
            .replace('十一', '11') \
            .replace('十', '10') \
            .replace('九', '9') \
            .replace('八', '8') \
            .replace('七', '7') \
            .replace('六', '6') \
            .replace('五', '5') \
            .replace('四', '4') \
            .replace('三', '3') \
            .replace('二', '2') \
            .replace('一', '1')

    @staticmethod
    def __str_wash(content):
        return str(content).strip().rstrip('\n')

    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master.title('收入业绩奖励统计')
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        # 设置界面宽度为530，高度为365像素，并且基于屏幕居中
        width = 700
        height = 450
        size = '%dx%d+%d+%d' % (width, height, (screenwidth - width) / 2, (screenheight - height) / 2)
        self.master.geometry(size)
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        Label(self, text="").grid(row=0, pady=5, columnspan=3)

        Label(self, text="输入文件：").grid(row=1, column=0, sticky='W', pady=5)
        self.inFile = StringVar()
        Entry(self, textvariable=self.inFile, width=40).grid(row=1, column=1, sticky='W')
        Button(self, text=" ... ", command=self.selectFile).grid(row=1, column=2, sticky='W', padx=2)

        Label(self, text="输出文件夹：").grid(row=2, column=0, sticky='W', pady=5)
        self.inDirectory = StringVar()
        Entry(self, textvariable=self.inDirectory, width=40).grid(row=2, column=1, sticky='W')
        Button(self, text=" ... ", command=self.selectPath).grid(row=2, column=2, sticky='W', padx=2)

        self.translateButton = Button(self, text="解析", command=self.translate, height=1, width=10)
        self.translateButton.grid(row=3, columnspan=3, padx=10, pady=10)

        self.run_log = ScrolledText(self, font=('楷体', 13), width=69, height=14)
        self.run_log.grid(row=4, columnspan=3, padx=20, pady=5, sticky='W')

    def selectPath(self):
        self.inDirectory.set(askdirectory())

    def selectFile(self):
        self.inFile.set(askopenfilename())

    def translate(self):
        try:
            month_fund_business = self.read_excel(self.inFile.get())
            self.write_excel(self.inDirectory.get(), month_fund_business)
        except Exception:
            self.run_log_self("error: {0}".format(traceback.format_exc()))

    def run_log_self(self, message):
        print(message)
        """ 实时更新日志，固定用法 """
        self.run_log.config(state=NORMAL)
        self.run_log.insert(END, "\n" + str(message) + "\n")
        self.run_log.see(END)
        self.run_log.update()
        self.run_log.config(state=DISABLED)


if __name__ == '__main__':
    app = Report()
    app.mainloop()
